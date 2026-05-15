# -*- coding: utf-8 -*-
"""
Storage Analyzer

Σαρώνει τους φακέλους χρηστών σε Windows, υπολογίζει το συνολικό μέγεθος
κάθε προφίλ και βρίσκει τα μεγαλύτερα αρχεία ανά χρήστη. Τα αποτελέσματα εξάγονται χειροκίνητα σε Excel όταν το επιλέξει ο διαχειριστής.

Περιλαμβάνει προαιρετική, χειροκίνητη και επιβεβαιωμένη διαγραφή επιλεγμένων
Windows user profiles μέσω Win32_UserProfile, ώστε να αφαιρούνται καθαρά
και οι αντίστοιχες εγγραφές profile/registry των Windows.

Χρήση:
  python storage_analyzer.py

Προαιρετικό EXE:
  pyinstaller StorageAnalyzer.spec
"""

from __future__ import annotations

import ctypes
import heapq
import os
import queue
import re
import stat
import sys
import threading
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Optional
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

APP_TITLE = "Storage Analyzer"
DEFAULT_SCAN_ROOT = r"C:\Users"
REPARSE_ATTRIBUTE = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
PROTECTED_PROFILE_NAMES = {
    "public",
    "default",
    "default user",
    "all users",
    "defaultuser0",
    "wdagutilityaccount",
}

# Dark UI palette.  No external UI dependency is used, so the EXE remains simple.
COLOR_BG = "#0B1120"
COLOR_PANEL = "#111827"
COLOR_PANEL_2 = "#182235"
COLOR_FIELD = "#0F172A"
COLOR_BORDER = "#334155"
COLOR_TEXT = "#E5E7EB"
COLOR_MUTED = "#94A3B8"
COLOR_ACCENT = "#2563EB"
COLOR_ACCENT_HOVER = "#3B82F6"
COLOR_ACCENT_ACTIVE = "#1D4ED8"
COLOR_DANGER = "#DC2626"
COLOR_DANGER_HOVER = "#EF4444"
COLOR_SUCCESS = "#16A34A"
COLOR_SUCCESS_HOVER = "#22C55E"
COLOR_WARNING = "#F59E0B"
COLOR_WARNING_HOVER = "#FBBF24"
COLOR_WARNING_ACTIVE = "#D97706"
COLOR_DISABLED = "#475569"
COLOR_TREE = "#0F172A"
COLOR_TREE_ALT = "#131D2F"

FONT_BASE = ("Segoe UI", 10)
FONT_SMALL = ("Segoe UI", 9)
FONT_TITLE = ("Segoe UI", 18, "bold")
FONT_SECTION = ("Segoe UI", 10, "bold")
FONT_BUTTON = ("Segoe UI", 10, "bold")


@dataclass
class LargeFile:
    size_bytes: int
    path: str
    modified_ts: float


@dataclass
class UserScanResult:
    user_name: str
    folder_path: str
    total_bytes: int = 0
    file_count: int = 0
    folder_count: int = 0
    error_count: int = 0
    skipped_reparse_count: int = 0
    large_files: list[LargeFile] = field(default_factory=list)
    elapsed_seconds: float = 0.0


@dataclass
class ProfileDeleteResult:
    user_name: str
    folder_path: str
    status: str
    message: str
    sid: str = ""
    loaded: Optional[bool] = None
    special: Optional[bool] = None
    timestamp: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


class ScanCancelled(Exception):
    pass


class RoundedButton(tk.Canvas):
    """Small dependency-free rounded button for Tkinter dark UI."""

    def __init__(
        self,
        master,
        text: str,
        command=None,
        width: int = 150,
        height: int = 36,
        radius: int = 16,
        bg_color: str = COLOR_ACCENT,
        hover_color: str = COLOR_ACCENT_HOVER,
        active_color: str = COLOR_ACCENT_ACTIVE,
        disabled_color: str = COLOR_DISABLED,
        text_color: str = "#FFFFFF",
        disabled_text_color: str = "#CBD5E1",
        outline_color: str = "",
        **kwargs,
    ):
        super().__init__(
            master,
            width=width,
            height=height,
            bg=kwargs.pop("background", COLOR_BG),
            highlightthickness=0,
            bd=0,
            relief="flat",
            cursor="hand2",
            **kwargs,
        )
        self._text = text
        self._command = command
        self._radius = radius
        self._normal_color = bg_color
        self._hover_color = hover_color
        self._active_color = active_color
        self._disabled_color = disabled_color
        self._text_color = text_color
        self._disabled_text_color = disabled_text_color
        self._outline_color = outline_color
        self._state = "normal"
        self._inside = False
        self._pressed = False

        self.bind("<Configure>", lambda _event: self._draw())
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<ButtonPress-1>", self._on_press)
        self.bind("<ButtonRelease-1>", self._on_release)
        self._draw()

    def configure(self, cnf=None, **kwargs):  # type: ignore[override]
        if isinstance(cnf, str):
            return super().configure(cnf)
        merged = {}
        if isinstance(cnf, dict):
            merged.update(cnf)
        merged.update(kwargs)

        if "state" in merged:
            self._state = str(merged.pop("state"))
            self.configure(cursor="arrow" if self._state == "disabled" else "hand2")
            self._draw()
        if "text" in merged:
            self._text = str(merged.pop("text"))
            self._draw()
        if "command" in merged:
            self._command = merged.pop("command")
        if "background" in merged:
            super().configure(bg=merged.pop("background"))
            self._draw()
        if "bg" in merged:
            super().configure(bg=merged.pop("bg"))
            self._draw()
        if merged:
            return super().configure(**merged)
        return None

    config = configure

    def cget(self, key):  # type: ignore[override]
        if key == "state":
            return self._state
        if key == "text":
            return self._text
        return super().cget(key)

    def _current_color(self) -> str:
        if self._state == "disabled":
            return self._disabled_color
        if self._pressed:
            return self._active_color
        if self._inside:
            return self._hover_color
        return self._normal_color

    def _draw(self) -> None:
        self.delete("all")
        width = max(2, int(self.winfo_width() or int(self["width"])))
        height = max(2, int(self.winfo_height() or int(self["height"])))
        radius = min(self._radius, height // 2, width // 2)
        x1, y1, x2, y2 = 1, 1, width - 1, height - 1
        fill = self._current_color()
        text_fill = self._disabled_text_color if self._state == "disabled" else self._text_color

        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1,
        ]
        self.create_polygon(points, smooth=True, fill=fill, outline=self._outline_color)
        self.create_text(
            width // 2,
            height // 2,
            text=self._text,
            fill=text_fill,
            font=FONT_BUTTON,
        )

    def _on_enter(self, _event) -> None:
        self._inside = True
        self._draw()

    def _on_leave(self, _event) -> None:
        self._inside = False
        self._pressed = False
        self._draw()

    def _on_press(self, _event) -> None:
        if self._state == "disabled":
            return
        self._pressed = True
        self._draw()

    def _on_release(self, _event) -> None:
        if self._state == "disabled":
            return
        was_pressed = self._pressed
        self._pressed = False
        self._draw()
        if was_pressed and self._inside and callable(self._command):
            self._command()


class NumberStepper(tk.Frame):
    """Visible plus/minus numeric control for the dark UI."""

    def __init__(
        self,
        master,
        variable: tk.IntVar,
        from_: int,
        to: int,
        step: int = 1,
        width: int = 7,
        background: str = COLOR_PANEL,
    ):
        super().__init__(master, bg=background, highlightthickness=0)
        self.variable = variable
        self.from_ = int(from_)
        self.to = int(to)
        self.step = int(step) if int(step) > 0 else 1

        self.minus_btn = RoundedButton(
            self,
            text="−",
            command=lambda: self._change(-self.step),
            width=36,
            height=32,
            radius=14,
            bg_color=COLOR_WARNING,
            hover_color=COLOR_WARNING_HOVER,
            active_color=COLOR_WARNING_ACTIVE,
            background=background,
        )
        self.minus_btn.pack(side="left")

        self.entry = tk.Entry(
            self,
            textvariable=self.variable,
            width=width,
            bg=COLOR_FIELD,
            fg=COLOR_TEXT,
            insertbackground=COLOR_TEXT,
            justify="center",
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground=COLOR_BORDER,
            highlightcolor=COLOR_WARNING,
            font=FONT_BASE,
        )
        self.entry.pack(side="left", padx=6, ipady=5)
        self.entry.bind("<Return>", lambda _event: self._normalize())
        self.entry.bind("<FocusOut>", lambda _event: self._normalize())

        self.plus_btn = RoundedButton(
            self,
            text="+",
            command=lambda: self._change(self.step),
            width=36,
            height=32,
            radius=14,
            bg_color=COLOR_WARNING,
            hover_color=COLOR_WARNING_HOVER,
            active_color=COLOR_WARNING_ACTIVE,
            background=background,
        )
        self.plus_btn.pack(side="left")
        self._normalize()

    def _current_value(self) -> int:
        try:
            return int(self.variable.get())
        except Exception:
            return self.from_

    def _set_value(self, value: int) -> None:
        value = max(self.from_, min(self.to, int(value)))
        self.variable.set(value)

    def _change(self, delta: int) -> None:
        self._set_value(self._current_value() + int(delta))

    def _normalize(self) -> None:
        self._set_value(self._current_value())

    def set_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        self.entry.configure(state=state)
        self.minus_btn.configure(state=state)
        self.plus_btn.configure(state=state)


def resource_path(relative_path: str) -> str:
    """Return a resource path that works from source and from a PyInstaller bundle."""
    try:
        base_path = Path(getattr(sys, "_MEIPASS"))  # type: ignore[attr-defined]
    except Exception:
        base_path = Path(__file__).resolve().parent
    return str(base_path / relative_path)


def is_windows() -> bool:
    return os.name == "nt"


def is_admin() -> bool:
    if not is_windows():
        return False
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False



def is_reparse_point(entry: os.DirEntry) -> bool:
    """Αποφεύγει junctions/symlinks για να μη διπλομετρήσει ή κολλήσει σε κύκλους."""
    try:
        if entry.is_symlink():
            return True
        attrs = getattr(entry.stat(follow_symlinks=False), "st_file_attributes", 0)
        return bool(attrs & REPARSE_ATTRIBUTE)
    except OSError:
        return False


def safe_modified_ts(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except OSError:
        return 0.0


def format_bytes(num: int) -> str:
    value = float(num)
    for unit in ("B", "KB", "MB", "GB", "TB", "PB"):
        if value < 1024.0 or unit == "PB":
            if unit == "B":
                return f"{int(value)} {unit}"
            return f"{value:.2f} {unit}"
        value /= 1024.0
    return f"{num} B"


def sanitize_filename_part(value: str) -> str:
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if ch in invalid else ch for ch in value)
    return cleaned.strip(" .") or "export"


def normalize_win_path(value: str) -> str:
    try:
        return os.path.normcase(os.path.normpath(os.path.abspath(value)))
    except Exception:
        return os.path.normcase(os.path.normpath(str(value)))


def natural_sort_key(value: str) -> tuple[Any, ...]:
    """Natural key so numeric usernames/registry numbers sort as numbers, not text."""
    parts = re.split(r"(\d+)", str(value).casefold())
    return tuple(int(part) if part.isdigit() else part for part in parts)


def is_current_user_profile_path(folder_path: str) -> bool:
    current_profile = os.environ.get("USERPROFILE", "")
    if not current_profile:
        return False
    return normalize_win_path(folder_path) == normalize_win_path(current_profile)


def is_protected_profile_folder(folder_path: str) -> bool:
    name = Path(folder_path).name.strip().lower()
    return name in PROTECTED_PROFILE_NAMES or is_current_user_profile_path(folder_path)


def _profile_bool(value: Any) -> Optional[bool]:
    """Convert WMI/COM boolean-like values to Python bool without guessing unknowns."""
    if value is None:
        return None
    try:
        return bool(value)
    except Exception:
        return None


def delete_windows_user_profile(user_name: str, folder_path: str) -> ProfileDeleteResult:
    """Delete a Windows user profile through Win32_UserProfile using Python COM/WMI.

    No PowerShell is used.  The function intentionally avoids plain folder
    deletion because that can leave ProfileList registry entries and broken
    Windows profile metadata behind.
    """
    if not is_windows():
        return ProfileDeleteResult(user_name, folder_path, "error", "Η διαγραφή προφίλ υποστηρίζεται μόνο σε Windows.")
    if not is_admin():
        return ProfileDeleteResult(user_name, folder_path, "error", "Απαιτούνται δικαιώματα Administrator.")
    if is_protected_profile_folder(folder_path):
        return ProfileDeleteResult(
            user_name,
            folder_path,
            "blocked",
            "Προστατευμένο ή τρέχον προφίλ. Δεν επιτρέπεται διαγραφή από το εργαλείο.",
        )

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
        from pywintypes import com_error  # type: ignore
    except Exception as exc:
        return ProfileDeleteResult(
            user_name,
            folder_path,
            "error",
            "Λείπει η βιβλιοθήκη pywin32 για native Python COM/WMI διαγραφή. "
            "Εγκατάσταση: pip install pywin32. "
            f"Λεπτομέρεια: {exc}",
        )

    target_path = normalize_win_path(folder_path)
    pythoncom.CoInitialize()
    try:
        locator = win32com.client.Dispatch("WbemScripting.SWbemLocator")
        services = locator.ConnectServer(".", "root\\cimv2")
        profiles = services.ExecQuery("SELECT SID, LocalPath, Loaded, Special FROM Win32_UserProfile")

        matched_profile = None
        for profile in profiles:
            local_path = str(getattr(profile, "LocalPath", "") or "").strip()
            if not local_path:
                continue
            if normalize_win_path(local_path) == target_path:
                matched_profile = profile
                break

        if matched_profile is None:
            return ProfileDeleteResult(
                user_name,
                folder_path,
                "not_found",
                "Δεν βρέθηκε Win32_UserProfile entry για αυτόν τον φάκελο. Δεν έγινε απλή διαγραφή φακέλου για λόγους ασφάλειας.",
            )

        sid = str(getattr(matched_profile, "SID", "") or "")
        loaded = _profile_bool(getattr(matched_profile, "Loaded", None))
        special = _profile_bool(getattr(matched_profile, "Special", None))

        if special:
            return ProfileDeleteResult(
                user_name,
                folder_path,
                "blocked",
                "Το προφίλ είναι Special/System profile και δεν διαγράφηκε.",
                sid=sid,
                loaded=loaded,
                special=special,
            )
        if loaded:
            return ProfileDeleteResult(
                user_name,
                folder_path,
                "blocked",
                "Το προφίλ είναι φορτωμένο/ενεργό. Κάνε log off τον χρήστη και ξαναδοκίμασε.",
                sid=sid,
                loaded=loaded,
                special=special,
            )

        try:
            matched_profile.Delete_()
        except com_error as exc:
            return ProfileDeleteResult(
                user_name,
                folder_path,
                "error",
                f"Απέτυχε η native WMI διαγραφή Win32_UserProfile: {exc}",
                sid=sid,
                loaded=loaded,
                special=special,
            )

        return ProfileDeleteResult(
            user_name,
            folder_path,
            "deleted",
            "Το προφίλ διαγράφηκε μέσω native Python COM/WMI Win32_UserProfile. Δεν χρησιμοποιήθηκε PowerShell.",
            sid=sid,
            loaded=loaded,
            special=special,
        )
    except Exception as exc:
        return ProfileDeleteResult(user_name, folder_path, "error", str(exc))
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

def iter_user_folders(root: str) -> Iterable[Path]:
    root_path = Path(root)
    if not root_path.exists():
        raise FileNotFoundError(f"Δεν βρέθηκε ο φάκελος: {root}")
    if not root_path.is_dir():
        raise NotADirectoryError(f"Δεν είναι φάκελος: {root}")

    with os.scandir(root) as it:
        for entry in it:
            try:
                if not entry.is_dir(follow_symlinks=False):
                    continue
                if is_reparse_point(entry):
                    continue
                yield Path(entry.path)
            except OSError:
                continue


def scan_one_user_folder(
    folder: Path,
    top_n_files: int,
    min_large_file_bytes: int,
    cancel_event: threading.Event,
    progress_callback=None,
) -> UserScanResult:
    start = time.perf_counter()
    result = UserScanResult(user_name=folder.name, folder_path=str(folder))
    large_heap: list[tuple[int, str, float]] = []
    stack = [str(folder)]
    last_progress = 0.0

    while stack:
        if cancel_event.is_set():
            raise ScanCancelled()

        current = stack.pop()
        try:
            with os.scandir(current) as it:
                for entry in it:
                    if cancel_event.is_set():
                        raise ScanCancelled()

                    try:
                        if is_reparse_point(entry):
                            result.skipped_reparse_count += 1
                            continue

                        if entry.is_dir(follow_symlinks=False):
                            result.folder_count += 1
                            stack.append(entry.path)
                            continue

                        if entry.is_file(follow_symlinks=False):
                            st = entry.stat(follow_symlinks=False)
                            size = int(st.st_size)
                            result.total_bytes += size
                            result.file_count += 1

                            if top_n_files > 0 and size >= min_large_file_bytes:
                                modified_ts = getattr(st, "st_mtime", safe_modified_ts(entry.path))
                                item = (size, entry.path, float(modified_ts))
                                if len(large_heap) < top_n_files:
                                    heapq.heappush(large_heap, item)
                                elif size > large_heap[0][0]:
                                    heapq.heapreplace(large_heap, item)

                    except PermissionError:
                        result.error_count += 1
                    except OSError:
                        result.error_count += 1

            now = time.perf_counter()
            if progress_callback and now - last_progress >= 0.4:
                last_progress = now
                progress_callback(result)

        except PermissionError:
            result.error_count += 1
        except OSError:
            result.error_count += 1

    result.large_files = [
        LargeFile(size_bytes=size, path=path, modified_ts=modified_ts)
        for size, path, modified_ts in sorted(large_heap, key=lambda x: x[0], reverse=True)
    ]
    result.elapsed_seconds = time.perf_counter() - start
    return result


def sorted_results(results: list[UserScanResult]) -> list[UserScanResult]:
    return sorted(results, key=lambda x: x.total_bytes, reverse=True)


def iter_large_file_rows(results: list[UserScanResult]):
    for r in sorted_results(results):
        for lf in r.large_files:
            modified = (
                datetime.fromtimestamp(lf.modified_ts).strftime("%Y-%m-%d %H:%M:%S")
                if lf.modified_ts
                else ""
            )
            yield r, lf, modified


def export_excel_results(results: list[UserScanResult], out: Path, timestamp: str, deletion_log: Optional[list[ProfileDeleteResult]] = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Λείπει η βιβλιοθήκη openpyxl. Κάνε εγκατάσταση με: pip install openpyxl"
        ) from exc

    xlsx_path = out / f"storage_analyzer_report_{timestamp}.xlsx"
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Report Info"
    ws_summary = wb.create_sheet("User Summary")
    ws_large = wb.create_sheet("Large Files")
    ws_delete = wb.create_sheet("Deletion Log") if deletion_log else None

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_side = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header(ws, row: int, max_col: int) -> None:
        for cell in ws[row][:max_col]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # Metadata
    meta_rows = [
        ("Report", APP_TITLE),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Users scanned", len(results)),
        ("Total size bytes", sum(r.total_bytes for r in results)),
        ("Total size", format_bytes(sum(r.total_bytes for r in results))),
        ("Total files", sum(r.file_count for r in results)),
        ("Total folders", sum(r.folder_count for r in results)),
        ("Total scan errors", sum(r.error_count for r in results)),
        ("Deletion log entries", len(deletion_log or [])),
        ("Note", "Το Excel δημιουργείται χειροκίνητα από τον διαχειριστή. Η διαγραφή προφίλ, αν έγινε, καταγράφεται στο Deletion Log."),
    ]
    ws_meta.append(["Field", "Value"])
    style_header(ws_meta, 1, 2)
    for row in meta_rows:
        ws_meta.append(row)
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 90
    for row in ws_meta.iter_rows(min_row=1, max_row=ws_meta.max_row, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws_meta.freeze_panes = "A2"

    # Summary sheet
    summary_headers = [
        "User",
        "Folder",
        "Total Bytes",
        "Total Size",
        "Files",
        "Folders",
        "Errors",
        "Skipped Reparse/Junctions",
        "Elapsed Seconds",
    ]
    ws_summary.append(summary_headers)
    style_header(ws_summary, 1, len(summary_headers))
    for r in sorted_results(results):
        ws_summary.append([
            r.user_name,
            r.folder_path,
            r.total_bytes,
            format_bytes(r.total_bytes),
            r.file_count,
            r.folder_count,
            r.error_count,
            r.skipped_reparse_count,
            round(r.elapsed_seconds, 2),
        ])
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions

    # Large files sheet
    large_headers = ["User", "File Size Bytes", "File Size", "Modified", "Path"]
    ws_large.append(large_headers)
    style_header(ws_large, 1, len(large_headers))
    for r, lf, modified in iter_large_file_rows(results):
        ws_large.append([r.user_name, lf.size_bytes, format_bytes(lf.size_bytes), modified, lf.path])
    ws_large.freeze_panes = "A2"
    ws_large.auto_filter.ref = ws_large.dimensions

    if ws_delete is not None:
        delete_headers = ["Timestamp", "User", "Folder", "Status", "Message", "SID", "Loaded", "Special"]
        ws_delete.append(delete_headers)
        style_header(ws_delete, 1, len(delete_headers))
        for entry in deletion_log or []:
            ws_delete.append([
                entry.timestamp,
                entry.user_name,
                entry.folder_path,
                entry.status,
                entry.message,
                entry.sid,
                "" if entry.loaded is None else str(entry.loaded),
                "" if entry.special is None else str(entry.special),
            ])
        ws_delete.freeze_panes = "A2"
        ws_delete.auto_filter.ref = ws_delete.dimensions

    # Formatting and widths
    sheet_widths = {
        "Report Info": {"A": 24, "B": 90},
        "User Summary": {"A": 24, "B": 65, "C": 18, "D": 14, "E": 14, "F": 14, "G": 12, "H": 24, "I": 16},
        "Large Files": {"A": 24, "B": 18, "C": 14, "D": 20, "E": 110},
        "Deletion Log": {"A": 20, "B": 24, "C": 70, "D": 14, "E": 90, "F": 46, "G": 12, "H": 12},
    }
    export_sheets = [ws_summary, ws_large]
    if ws_delete is not None:
        export_sheets.append(ws_delete)
    for ws in export_sheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = subheader_fill
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = sheet_widths.get(ws.title, {}).get(letter, 18)
        for row_idx in range(1, min(ws.max_row, 5000) + 1):
            ws.row_dimensions[row_idx].height = 18

    # Numeric formats
    for ws in export_sheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, int):
                    cell.number_format = '#,##0'
                elif isinstance(cell.value, float):
                    cell.number_format = '0.00'

    wb.save(xlsx_path)
    return xlsx_path


def export_results(results: list[UserScanResult], output_dir: str, deletion_log: Optional[list[ProfileDeleteResult]] = None) -> dict[str, str]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    xlsx_path = export_excel_results(results, out, timestamp, deletion_log=deletion_log)

    return {
        "excel": str(xlsx_path),
    }


class StorageAnalyzerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1220x800")
        self.minsize(1020, 650)
        self.configure(bg=COLOR_BG)
        self._set_window_icon()

        self.result_queue: queue.Queue = queue.Queue()
        self.cancel_event = threading.Event()
        self.worker: Optional[threading.Thread] = None
        self.delete_worker: Optional[threading.Thread] = None
        self.results: list[UserScanResult] = []
        self.current_partial: dict[str, UserScanResult] = {}
        self.delete_selection: set[str] = set()
        self.deletion_log: list[ProfileDeleteResult] = []
        self.last_exported: dict[str, str] = {}
        self.summary_sort_column = "bytes"
        self.summary_sort_reverse = True
        self.files_sort_column = "bytes"
        self.files_sort_reverse = True
        self._last_delete_toggle_path = ""
        self._last_delete_toggle_ts = 0.0

        self.scan_root_var = tk.StringVar(value=DEFAULT_SCAN_ROOT)
        self.output_dir_var = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.top_n_var = tk.IntVar(value=20)
        self.min_mb_var = tk.IntVar(value=50)
        self.status_var = tk.StringVar(value="Έτοιμο.")
        self.admin_var = tk.StringVar(value="ADMIN MODE")

        self._configure_dark_theme()
        self._build_ui()
        self.after(200, self._poll_queue)


    def _set_window_icon(self) -> None:
        """Set the application window icon when app.ico is available."""
        icon_path = resource_path("app.ico")
        try:
            if is_windows():
                self.iconbitmap(icon_path)
            else:
                icon_img = tk.PhotoImage(file=icon_path)
                self.iconphoto(True, icon_img)
                self._icon_img_ref = icon_img
        except Exception:
            # Icon failure must never block the storage audit tool.
            pass

    def _configure_dark_theme(self) -> None:
        self.style = ttk.Style(self)
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        self.style.configure(".", font=FONT_BASE)
        self.style.configure("Dark.TFrame", background=COLOR_BG)
        self.style.configure("Panel.TFrame", background=COLOR_PANEL)
        self.style.configure("SoftPanel.TFrame", background=COLOR_PANEL_2)
        self.style.configure("Dark.TLabel", background=COLOR_BG, foreground=COLOR_TEXT)
        self.style.configure("Panel.TLabel", background=COLOR_PANEL, foreground=COLOR_TEXT)
        self.style.configure("Muted.TLabel", background=COLOR_BG, foreground=COLOR_MUTED, font=FONT_SMALL)
        self.style.configure("PanelMuted.TLabel", background=COLOR_PANEL, foreground=COLOR_MUTED, font=FONT_SMALL)
        self.style.configure("Section.TLabel", background=COLOR_PANEL, foreground=COLOR_TEXT, font=FONT_SECTION)

        self.style.configure(
            "Card.TLabelframe",
            background=COLOR_PANEL,
            foreground=COLOR_TEXT,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            relief="solid",
        )
        self.style.configure(
            "Card.TLabelframe.Label",
            background=COLOR_PANEL,
            foreground=COLOR_TEXT,
            font=FONT_SECTION,
        )
        self.style.configure(
            "Dark.TEntry",
            fieldbackground=COLOR_FIELD,
            background=COLOR_FIELD,
            foreground=COLOR_TEXT,
            insertcolor=COLOR_TEXT,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            padding=7,
        )
        self.style.map(
            "Dark.TEntry",
            fieldbackground=[("readonly", COLOR_FIELD), ("disabled", COLOR_PANEL_2)],
            foreground=[("disabled", COLOR_MUTED)],
        )
        self.style.configure(
            "Dark.TSpinbox",
            fieldbackground=COLOR_FIELD,
            background=COLOR_FIELD,
            foreground=COLOR_TEXT,
            arrowsize=14,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            padding=5,
        )
        self.style.configure(
            "Treeview",
            background=COLOR_TREE,
            fieldbackground=COLOR_TREE,
            foreground=COLOR_TEXT,
            bordercolor=COLOR_BORDER,
            lightcolor=COLOR_BORDER,
            darkcolor=COLOR_BORDER,
            rowheight=28,
            font=FONT_BASE,
        )
        self.style.configure(
            "Treeview.Heading",
            background=COLOR_PANEL_2,
            foreground=COLOR_TEXT,
            font=FONT_SECTION,
            relief="flat",
            padding=8,
            anchor="center",
        )
        self.style.map(
            "Treeview",
            background=[("selected", COLOR_ACCENT)],
            foreground=[("selected", "#FFFFFF")],
        )
        self.style.map(
            "Treeview.Heading",
            background=[("active", COLOR_ACCENT)],
            foreground=[("active", "#FFFFFF")],
        )
        self.style.configure(
            "Vertical.TScrollbar",
            background=COLOR_PANEL_2,
            troughcolor=COLOR_BG,
            bordercolor=COLOR_BG,
            arrowcolor=COLOR_TEXT,
        )
        self.style.configure("TPanedwindow", background=COLOR_BG)

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=14, style="Dark.TFrame")
        root.pack(fill="both", expand=True)

        header = tk.Frame(root, bg=COLOR_BG, highlightthickness=0)
        header.pack(fill="x", pady=(0, 12))

        title_box = tk.Frame(header, bg=COLOR_BG)
        title_box.pack(side="left", fill="x", expand=True)
        tk.Label(
            title_box,
            text=APP_TITLE,
            bg=COLOR_BG,
            fg=COLOR_TEXT,
            font=FONT_TITLE,
        ).pack(anchor="w")
        tk.Label(
            title_box,
            text="Σάρωση προφίλ χρηστών, μεγάλα αρχεία και ασφαλής διαγραφή επιλεγμένων Windows profiles.",
            bg=COLOR_BG,
            fg=COLOR_MUTED,
            font=FONT_SMALL,
        ).pack(anchor="w", pady=(2, 0))

        admin_pill_color = COLOR_SUCCESS
        self.admin_pill = tk.Label(
            header,
            textvariable=self.admin_var,
            bg=admin_pill_color,
            fg="#FFFFFF",
            padx=14,
            pady=7,
            font=("Segoe UI", 9, "bold"),
        )
        self.admin_pill.pack(side="right", padx=(12, 0))

        top = ttk.LabelFrame(root, text="Ρυθμίσεις σάρωσης", padding=14, style="Card.TLabelframe")
        top.pack(fill="x")

        ttk.Label(top, text="Φάκελος χρηστών", style="Panel.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(top, textvariable=self.scan_root_var, style="Dark.TEntry").grid(row=0, column=1, sticky="ew", pady=6)
        RoundedButton(
            top,
            text="Επιλογή...",
            command=self._choose_scan_root,
            width=118,
            height=34,
            bg_color=COLOR_PANEL_2,
            hover_color=COLOR_ACCENT,
            active_color=COLOR_ACCENT_ACTIVE,
            outline_color=COLOR_BORDER,
            background=COLOR_PANEL,
        ).grid(row=0, column=2, padx=(10, 0), pady=6)

        ttk.Label(top, text="Φάκελος export", style="Panel.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(top, textvariable=self.output_dir_var, style="Dark.TEntry").grid(row=1, column=1, sticky="ew", pady=6)
        RoundedButton(
            top,
            text="Επιλογή...",
            command=self._choose_output_dir,
            width=118,
            height=34,
            bg_color=COLOR_PANEL_2,
            hover_color=COLOR_ACCENT,
            active_color=COLOR_ACCENT_ACTIVE,
            outline_color=COLOR_BORDER,
            background=COLOR_PANEL,
        ).grid(row=1, column=2, padx=(10, 0), pady=6)

        ttk.Label(top, text="Μεγάλα αρχεία / χρήστη", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        options = ttk.Frame(top, style="Panel.TFrame")
        options.grid(row=2, column=1, sticky="w", pady=6)
        NumberStepper(
            options,
            variable=self.top_n_var,
            from_=1,
            to=500,
            step=1,
            width=6,
            background=COLOR_PANEL,
        ).pack(side="left")
        ttk.Label(options, text="  Ελάχιστο μέγεθος MB", style="Panel.TLabel").pack(side="left", padx=(18, 8))
        NumberStepper(
            options,
            variable=self.min_mb_var,
            from_=0,
            to=102400,
            step=1,
            width=7,
            background=COLOR_PANEL,
        ).pack(side="left")
        ttk.Label(
            top,
            text="Απαιτείται εκτέλεση ως Administrator. Διαφορετικά το πρόγραμμα δεν ξεκινά.",
            style="PanelMuted.TLabel",
        ).grid(row=3, column=1, sticky="w", pady=(2, 0))

        top.columnconfigure(1, weight=1)

        buttons = ttk.Frame(root, padding=(0, 12, 0, 10), style="Dark.TFrame")
        buttons.pack(fill="x")
        self.start_btn = RoundedButton(
            buttons,
            text="Έναρξη σάρωσης",
            command=self._start_scan,
            width=160,
            height=38,
            bg_color=COLOR_SUCCESS,
            hover_color=COLOR_SUCCESS_HOVER,
            active_color="#15803D",
            background=COLOR_BG,
        )
        self.start_btn.pack(side="left")
        self.cancel_btn = RoundedButton(
            buttons,
            text="Ακύρωση",
            command=self._cancel_scan,
            width=115,
            height=38,
            bg_color=COLOR_DANGER,
            hover_color=COLOR_DANGER_HOVER,
            active_color="#B91C1C",
            background=COLOR_BG,
        )
        self.cancel_btn.pack(side="left", padx=(8, 0))
        self.cancel_btn.configure(state="disabled")

        self.export_btn = RoundedButton(
            buttons,
            text="Export Excel",
            command=self._export,
            width=150,
            height=38,
            bg_color=COLOR_ACCENT,
            hover_color=COLOR_ACCENT_HOVER,
            active_color=COLOR_ACCENT_ACTIVE,
            background=COLOR_BG,
        )
        self.export_btn.pack(side="left", padx=(8, 0))
        self.export_btn.configure(state="disabled")

        self.delete_btn = RoundedButton(
            buttons,
            text="Διαγραφή επιλεγμένων",
            command=self._delete_selected_profiles,
            width=190,
            height=38,
            bg_color=COLOR_DANGER,
            hover_color=COLOR_DANGER_HOVER,
            active_color="#B91C1C",
            background=COLOR_BG,
        )
        self.delete_btn.pack(side="left", padx=(8, 0))
        self.delete_btn.configure(state="disabled")


        main_pane = ttk.PanedWindow(root, orient="vertical", style="TPanedwindow")
        main_pane.pack(fill="both", expand=True)

        summary_frame = ttk.LabelFrame(main_pane, text="Μέγεθος φακέλου ανά χρήστη", padding=10, style="Card.TLabelframe")
        self.summary_tree = ttk.Treeview(
            summary_frame,
            columns=("delete", "user", "size", "bytes", "files", "folders", "errors", "skipped", "path"),
            show="headings",
            height=10,
        )
        self._configure_sortable_heading(self.summary_tree, "delete", "Διαγραφή", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "user", "User / Username", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "size", "Μέγεθος", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "bytes", "Bytes", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "files", "Αρχεία", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "folders", "Φάκελοι", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "errors", "Σφάλματα", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "skipped", "Junctions/Symlinks", self._sort_summary_by)
        self._configure_sortable_heading(self.summary_tree, "path", "Φάκελος", self._sort_summary_by)
        self.summary_tree.column("delete", width=90, anchor="center")
        self.summary_tree.column("user", width=150, anchor="center")
        self.summary_tree.column("size", width=110, anchor="center")
        self.summary_tree.column("bytes", width=130, anchor="center")
        self.summary_tree.column("files", width=90, anchor="center")
        self.summary_tree.column("folders", width=90, anchor="center")
        self.summary_tree.column("errors", width=90, anchor="center")
        self.summary_tree.column("skipped", width=145, anchor="center")
        self.summary_tree.column("path", width=520, anchor="center")
        self.summary_tree.tag_configure("odd", background=COLOR_TREE)
        self.summary_tree.tag_configure("even", background=COLOR_TREE_ALT)
        self.summary_tree.pack(side="left", fill="both", expand=True)
        summary_scroll = ttk.Scrollbar(summary_frame, orient="vertical", command=self.summary_tree.yview, style="Vertical.TScrollbar")
        summary_scroll.pack(side="right", fill="y")
        self.summary_tree.configure(yscrollcommand=summary_scroll.set)
        self.summary_tree.bind("<<TreeviewSelect>>", self._on_user_selected)
        self.summary_tree.bind("<ButtonPress-1>", self._on_summary_click)
        main_pane.add(summary_frame, weight=1)

        files_frame = ttk.LabelFrame(main_pane, text="Μεγαλύτερα αρχεία επιλεγμένου χρήστη", padding=10, style="Card.TLabelframe")
        self.files_tree = ttk.Treeview(
            files_frame,
            columns=("size", "bytes", "modified", "path"),
            show="headings",
            height=12,
        )
        self._configure_sortable_heading(self.files_tree, "size", "Μέγεθος", self._sort_files_by)
        self._configure_sortable_heading(self.files_tree, "bytes", "Bytes", self._sort_files_by)
        self._configure_sortable_heading(self.files_tree, "modified", "Τροποποίηση", self._sort_files_by)
        self._configure_sortable_heading(self.files_tree, "path", "Αρχείο", self._sort_files_by)
        self.files_tree.column("size", width=120, anchor="center")
        self.files_tree.column("bytes", width=130, anchor="center")
        self.files_tree.column("modified", width=170, anchor="center")
        self.files_tree.column("path", width=760, anchor="center")
        self.files_tree.tag_configure("odd", background=COLOR_TREE)
        self.files_tree.tag_configure("even", background=COLOR_TREE_ALT)
        self.files_tree.pack(side="left", fill="both", expand=True)
        files_scroll = ttk.Scrollbar(files_frame, orient="vertical", command=self.files_tree.yview, style="Vertical.TScrollbar")
        files_scroll.pack(side="right", fill="y")
        self.files_tree.configure(yscrollcommand=files_scroll.set)
        main_pane.add(files_frame, weight=1)

        delete_results_frame = ttk.LabelFrame(root, text="Αποτελέσματα διαγραφής", padding=10, style="Card.TLabelframe")
        delete_results_frame.pack(fill="x", pady=(10, 0))
        self.delete_results_tree = ttk.Treeview(
            delete_results_frame,
            columns=("time", "user", "status", "message", "sid"),
            show="headings",
            height=5,
        )
        self._configure_sortable_heading(self.delete_results_tree, "time", "Ώρα", lambda col: None)
        self._configure_sortable_heading(self.delete_results_tree, "user", "User", lambda col: None)
        self._configure_sortable_heading(self.delete_results_tree, "status", "Status", lambda col: None)
        self._configure_sortable_heading(self.delete_results_tree, "message", "Μήνυμα", lambda col: None)
        self._configure_sortable_heading(self.delete_results_tree, "sid", "SID", lambda col: None)
        self.delete_results_tree.column("time", width=145, anchor="center")
        self.delete_results_tree.column("user", width=130, anchor="center")
        self.delete_results_tree.column("status", width=95, anchor="center")
        self.delete_results_tree.column("message", width=620, anchor="center")
        self.delete_results_tree.column("sid", width=230, anchor="center")
        self.delete_results_tree.tag_configure("deleted", background="#0F2E22")
        self.delete_results_tree.tag_configure("blocked", background="#33250B")
        self.delete_results_tree.tag_configure("failed", background="#3A1111")
        self.delete_results_tree.pack(side="left", fill="x", expand=True)
        delete_results_scroll = ttk.Scrollbar(delete_results_frame, orient="vertical", command=self.delete_results_tree.yview, style="Vertical.TScrollbar")
        delete_results_scroll.pack(side="right", fill="y")
        self.delete_results_tree.configure(yscrollcommand=delete_results_scroll.set)

        status_bar = tk.Frame(root, bg=COLOR_PANEL, highlightthickness=1, highlightbackground=COLOR_BORDER)
        status_bar.pack(fill="x", pady=(10, 0))
        tk.Label(
            status_bar,
            textvariable=self.status_var,
            bg=COLOR_PANEL,
            fg=COLOR_TEXT,
            padx=12,
            pady=8,
            anchor="w",
            font=FONT_SMALL,
        ).pack(side="left", fill="x", expand=True)



    def _configure_sortable_heading(self, tree: ttk.Treeview, column: str, text: str, callback: Callable[[str], None]) -> None:
        tree.heading(column, text=text, anchor="center", command=lambda c=column: callback(c))

    def _sort_summary_by(self, column: str) -> None:
        if column == "delete":
            return
        if self.summary_sort_column == column:
            self.summary_sort_reverse = not self.summary_sort_reverse
        else:
            self.summary_sort_column = column
            self.summary_sort_reverse = column in {"size", "bytes", "files", "folders", "errors", "skipped"}
        self._refresh_summary_tree()

    def _sort_files_by(self, column: str) -> None:
        if self.files_sort_column == column:
            self.files_sort_reverse = not self.files_sort_reverse
        else:
            self.files_sort_column = column
            self.files_sort_reverse = column in {"size", "bytes", "modified"}
        self._refresh_files_tree()

    def _summary_sort_key(self, result: UserScanResult) -> Any:
        column = self.summary_sort_column
        if column in {"size", "bytes"}:
            return result.total_bytes
        if column == "files":
            return result.file_count
        if column == "folders":
            return result.folder_count
        if column == "errors":
            return result.error_count
        if column == "skipped":
            return result.skipped_reparse_count
        if column == "path":
            return natural_sort_key(result.folder_path)
        return natural_sort_key(result.user_name)

    def _file_sort_key(self, large_file: LargeFile) -> Any:
        column = self.files_sort_column
        if column in {"size", "bytes"}:
            return large_file.size_bytes
        if column == "modified":
            return large_file.modified_ts or 0
        return natural_sort_key(large_file.path)

    def _selected_user_result(self) -> Optional[UserScanResult]:
        selection = self.summary_tree.selection()
        if not selection:
            return None
        user_name = selection[0]
        return next((r for r in self.results if r.user_name == user_name), None)


    def _choose_scan_root(self) -> None:
        selected = filedialog.askdirectory(title="Επιλογή φακέλου χρηστών", initialdir=self.scan_root_var.get())
        if selected:
            self.scan_root_var.set(selected)

    def _choose_output_dir(self) -> None:
        selected = filedialog.askdirectory(title="Επιλογή φακέλου export", initialdir=self.output_dir_var.get())
        if selected:
            self.output_dir_var.set(selected)

    def _clear_trees(self) -> None:
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        for item in self.delete_results_tree.get_children():
            self.delete_results_tree.delete(item)

    def _start_scan(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Σάρωση", "Η σάρωση ήδη εκτελείται.")
            return

        scan_root = self.scan_root_var.get().strip()
        if not scan_root:
            messagebox.showerror("Σφάλμα", "Δώσε φάκελο σάρωσης.")
            return

        try:
            top_n = int(self.top_n_var.get())
            min_mb = int(self.min_mb_var.get())
        except Exception:
            messagebox.showerror("Σφάλμα", "Τα πεδία αριθμών δεν είναι σωστά.")
            return

        if top_n < 1:
            messagebox.showerror("Σφάλμα", "Τα μεγάλα αρχεία ανά χρήστη πρέπει να είναι τουλάχιστον 1.")
            return
        if min_mb < 0:
            messagebox.showerror("Σφάλμα", "Το ελάχιστο μέγεθος MB δεν μπορεί να είναι αρνητικό.")
            return

        self.cancel_event.clear()
        self.results = []
        self.current_partial = {}
        self.delete_selection.clear()
        self.last_exported = {}
        self._clear_trees()
        self.start_btn.configure(state="disabled")
        self.cancel_btn.configure(state="normal")
        self.export_btn.configure(state="disabled")
        self.delete_btn.configure(state="disabled")
        self.status_var.set("Ξεκίνησε η σάρωση...")

        self.worker = threading.Thread(
            target=self._scan_worker,
            args=(scan_root, top_n, min_mb * 1024 * 1024),
            daemon=True,
        )
        self.worker.start()

    def _cancel_scan(self) -> None:
        self.cancel_event.set()
        self.status_var.set("Ζητήθηκε ακύρωση. Περιμένω να σταματήσει καθαρά η τρέχουσα εργασία...")

    def _scan_worker(self, scan_root: str, top_n: int, min_large_file_bytes: int) -> None:
        try:
            folders = list(iter_user_folders(scan_root))
            self.result_queue.put(("info", f"Βρέθηκαν {len(folders)} φάκελοι χρηστών."))

            for index, folder in enumerate(folders, start=1):
                if self.cancel_event.is_set():
                    raise ScanCancelled()
                self.result_queue.put(("info", f"Σάρωση {index}/{len(folders)}: {folder.name}"))

                def progress(partial_result: UserScanResult) -> None:
                    self.result_queue.put(("partial", partial_result))

                result = scan_one_user_folder(
                    folder=folder,
                    top_n_files=top_n,
                    min_large_file_bytes=min_large_file_bytes,
                    cancel_event=self.cancel_event,
                    progress_callback=progress,
                )
                self.result_queue.put(("result", result))

            self.result_queue.put(("done", None))
        except ScanCancelled:
            self.result_queue.put(("cancelled", None))
        except Exception as exc:
            self.result_queue.put(("error", str(exc)))

    def _poll_queue(self) -> None:
        try:
            while True:
                msg_type, payload = self.result_queue.get_nowait()

                if msg_type == "info":
                    self.status_var.set(str(payload))
                elif msg_type == "partial":
                    self._update_partial(payload)
                elif msg_type == "result":
                    self._add_result(payload)
                elif msg_type == "done":
                    self._finish_scan(cancelled=False)
                elif msg_type == "cancelled":
                    self._finish_scan(cancelled=True)
                elif msg_type == "delete_info":
                    self.status_var.set(str(payload))
                elif msg_type == "delete_result":
                    self._handle_delete_result(payload)
                elif msg_type == "delete_done":
                    self._finish_delete()
                elif msg_type == "delete_error":
                    self._finish_delete(error=str(payload))
                elif msg_type == "error":
                    self._finish_scan(cancelled=True)
                    messagebox.showerror("Σφάλμα σάρωσης", str(payload))
        except queue.Empty:
            pass
        self.after(200, self._poll_queue)

    def _update_partial(self, result: UserScanResult) -> None:
        self.current_partial[result.user_name] = result
        self.status_var.set(
            f"Σαρώνεται: {result.user_name} | "
            f"{format_bytes(result.total_bytes)} | "
            f"αρχεία: {result.file_count:,} | σφάλματα: {result.error_count:,}"
        )

    def _add_result(self, result: UserScanResult) -> None:
        self.results.append(result)
        self._refresh_summary_tree()
        self.status_var.set(
            f"Ολοκληρώθηκε: {result.user_name} | {format_bytes(result.total_bytes)} | "
            f"αρχεία: {result.file_count:,} | σφάλματα: {result.error_count:,}"
        )

    def _refresh_summary_tree(self) -> None:
        selected_user = None
        selection = self.summary_tree.selection()
        if selection:
            selected_user = selection[0]

        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)

        for index, r in enumerate(sorted(self.results, key=self._summary_sort_key, reverse=self.summary_sort_reverse)):
            tag = "even" if index % 2 == 0 else "odd"
            checked = "☑" if r.folder_path in self.delete_selection else "☐"
            self.summary_tree.insert(
                "",
                "end",
                iid=r.user_name,
                tags=(tag,),
                values=(
                    checked,
                    r.user_name,
                    format_bytes(r.total_bytes),
                    f"{r.total_bytes:,}",
                    f"{r.file_count:,}",
                    f"{r.folder_count:,}",
                    f"{r.error_count:,}",
                    f"{r.skipped_reparse_count:,}",
                    r.folder_path,
                ),
            )

        if selected_user and self.summary_tree.exists(selected_user):
            self.summary_tree.selection_set(selected_user)

    def _on_summary_click(self, event) -> None:
        region = self.summary_tree.identify("region", event.x, event.y)
        column = self.summary_tree.identify_column(event.x)
        item = self.summary_tree.identify_row(event.y)
        if region == "cell" and column == "#1" and item:
            result = next((r for r in self.results if r.user_name == item), None)
            if not result:
                return "break"

            # Some touchpads/mice can emit a quick second click event. Without a
            # small guard, the profile checkbox may toggle on and immediately off.
            now = time.monotonic()
            if (
                self._last_delete_toggle_path == result.folder_path
                and now - self._last_delete_toggle_ts < 0.35
            ):
                return "break"
            self._last_delete_toggle_path = result.folder_path
            self._last_delete_toggle_ts = now

            if result.folder_path in self.delete_selection:
                self.delete_selection.remove(result.folder_path)
            else:
                self.delete_selection.add(result.folder_path)
            self._refresh_summary_tree()
            if self.summary_tree.exists(item):
                self.summary_tree.focus(item)
                self.summary_tree.selection_set(item)
            self._update_delete_button_state()
            return "break"
        return None

    def _update_delete_button_state(self) -> None:
        scanning = bool(self.worker and self.worker.is_alive())
        deleting = bool(self.delete_worker and self.delete_worker.is_alive())
        enabled = bool(self.delete_selection) and not scanning and not deleting
        self.delete_btn.configure(state="normal" if enabled else "disabled")

    def _selected_delete_results(self) -> list[UserScanResult]:
        selected_paths = set(self.delete_selection)
        return [r for r in self.results if r.folder_path in selected_paths]

    def _delete_selected_profiles(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Διαγραφή", "Δεν μπορεί να γίνει διαγραφή όσο εκτελείται σάρωση.")
            return
        if self.delete_worker and self.delete_worker.is_alive():
            messagebox.showinfo("Διαγραφή", "Η διαγραφή ήδη εκτελείται.")
            return

        selected = self._selected_delete_results()
        if not selected:
            messagebox.showinfo("Διαγραφή", "Δεν έχεις επιλέξει προφίλ για διαγραφή.")
            return

        protected = [r for r in selected if is_protected_profile_folder(r.folder_path)]
        allowed = [r for r in selected if not is_protected_profile_folder(r.folder_path)]
        if protected:
            protected_text = "\n".join(f"- {r.user_name}: {r.folder_path}" for r in protected[:10])
            messagebox.showwarning(
                "Προστατευμένα προφίλ",
                "Τα παρακάτω προφίλ δεν θα διαγραφούν γιατί είναι προστατευμένα ή είναι το τρέχον προφίλ:\n\n"
                f"{protected_text}"
                + ("\n..." if len(protected) > 10 else ""),
            )
        if not allowed:
            self.delete_selection.clear()
            self._refresh_summary_tree()
            self._update_delete_button_state()
            return

        preview = "\n".join(f"- {r.user_name}: {r.folder_path}" for r in allowed[:12])
        if len(allowed) > 12:
            preview += f"\n...και άλλα {len(allowed) - 12}"

        if not messagebox.askyesno(
            "Επιβεβαίωση διαγραφής προφίλ",
            "ΠΡΟΣΟΧΗ: Η ενέργεια είναι μόνιμη. Θα διαγραφούν τα επιλεγμένα Windows user profiles "
            "μέσω native Python COM/WMI Win32_UserProfile, μαζί με τα αντίστοιχα Windows profile metadata/registry entries. Δεν θα χρησιμοποιηθεί PowerShell.\n\n"
            f"Προφίλ προς διαγραφή: {len(allowed)}\n\n{preview}\n\nΣυνέχεια;",
            icon="warning",
        ):
            return

        typed = simpledialog.askstring(
            "Τελική επιβεβαίωση",
            "Για να συνεχίσει η διαγραφή, πληκτρολόγησε ακριβώς:\n\nDELETE",
            parent=self,
        )
        if typed != "DELETE":
            messagebox.showinfo("Ακύρωση", "Η διαγραφή ακυρώθηκε. Δεν έγινε καμία αλλαγή.")
            return

        self.start_btn.configure(state="disabled")
        self.cancel_btn.configure(state="disabled")
        self.export_btn.configure(state="disabled")
        self.delete_btn.configure(state="disabled")
        self.status_var.set(f"Ξεκίνησε διαγραφή {len(allowed)} προφίλ...")

        self.delete_worker = threading.Thread(target=self._delete_profiles_worker, args=(allowed,), daemon=True)
        self.delete_worker.start()

    def _delete_profiles_worker(self, selected: list[UserScanResult]) -> None:
        try:
            for index, result in enumerate(selected, start=1):
                self.result_queue.put(("delete_info", f"Διαγραφή {index}/{len(selected)}: {result.user_name}"))
                delete_result = delete_windows_user_profile(result.user_name, result.folder_path)
                self.result_queue.put(("delete_result", delete_result))
            self.result_queue.put(("delete_done", None))
        except Exception as exc:
            self.result_queue.put(("delete_error", str(exc)))

    def _handle_delete_result(self, result: ProfileDeleteResult) -> None:
        self.deletion_log.append(result)
        self._append_delete_result_panel(result)
        if result.status == "deleted":
            self.results = [r for r in self.results if r.folder_path != result.folder_path]
            self.delete_selection.discard(result.folder_path)
            self.status_var.set(f"Διαγράφηκε: {result.user_name}")
        else:
            self.delete_selection.discard(result.folder_path)
            self.status_var.set(f"Δεν διαγράφηκε: {result.user_name} | {result.status}: {result.message}")
        self._refresh_summary_tree()
        self._update_delete_button_state()

    def _finish_delete(self, error: str = "") -> None:
        self.start_btn.configure(state="normal")
        self.cancel_btn.configure(state="disabled")
        self.export_btn.configure(state="normal" if (self.results or self.deletion_log) else "disabled")
        self._update_delete_button_state()
        deleted = sum(1 for r in self.deletion_log if r.status == "deleted")
        blocked = sum(1 for r in self.deletion_log if r.status == "blocked")
        failed = sum(1 for r in self.deletion_log if r.status in {"error", "not_found"})
        if error:
            self.status_var.set("Η διαγραφή σταμάτησε με σφάλμα.")
            messagebox.showerror("Σφάλμα διαγραφής", error)
            return
        self.status_var.set(
            f"Η διαγραφή ολοκληρώθηκε. Διαγράφηκαν: {deleted} | Μπλοκαρίστηκαν: {blocked} | Απέτυχαν/δεν βρέθηκαν: {failed}"
        )
        messagebox.showinfo(
            "Η διαγραφή ολοκληρώθηκε",
            "Η διαδικασία διαγραφής ολοκληρώθηκε.\n\n"
            f"Διαγράφηκαν: {deleted}\n"
            f"Μπλοκαρίστηκαν: {blocked}\n"
            f"Απέτυχαν/δεν βρέθηκαν: {failed}\n\n"
            "Πάτησε Export Excel αν θέλεις αναφορά με το Deletion Log.",
        )

    def _on_user_selected(self, _event=None) -> None:
        self._refresh_files_tree()

    def _refresh_files_tree(self) -> None:
        result = self._selected_user_result()
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        if not result:
            return

        for index, lf in enumerate(sorted(result.large_files, key=self._file_sort_key, reverse=self.files_sort_reverse)):
            modified = (
                datetime.fromtimestamp(lf.modified_ts).strftime("%Y-%m-%d %H:%M:%S")
                if lf.modified_ts
                else ""
            )
            tag = "even" if index % 2 == 0 else "odd"
            self.files_tree.insert(
                "",
                "end",
                tags=(tag,),
                values=(format_bytes(lf.size_bytes), f"{lf.size_bytes:,}", modified, lf.path),
            )

    def _append_delete_result_panel(self, result: ProfileDeleteResult) -> None:
        if result.status == "deleted":
            tag = "deleted"
        elif result.status == "blocked":
            tag = "blocked"
        else:
            tag = "failed"
        self.delete_results_tree.insert(
            "",
            "end",
            tags=(tag,),
            values=(result.timestamp, result.user_name, result.status, result.message, result.sid),
        )
        children = self.delete_results_tree.get_children()
        if children:
            self.delete_results_tree.see(children[-1])

    def _finish_scan(self, cancelled: bool) -> None:
        self.start_btn.configure(state="normal")
        self.cancel_btn.configure(state="disabled")
        self.export_btn.configure(state="normal" if self.results else "disabled")
        self._update_delete_button_state()
        if cancelled:
            self.status_var.set(f"Η σάρωση ακυρώθηκε. Ολοκληρωμένοι χρήστες: {len(self.results)}")
        else:
            total = sum(r.total_bytes for r in self.results)
            completion_message = (
                f"Η σάρωση ολοκληρώθηκε. Χρήστες: {len(self.results)} | "
                f"Συνολικό μέγεθος: {format_bytes(total)} | Πάτησε Export Excel για αναφορά."
            )
            self.status_var.set(completion_message)
            messagebox.showinfo(
                "Η σάρωση ολοκληρώθηκε",
                "Η σάρωση ολοκληρώθηκε επιτυχώς.\n\n"
                f"Χρήστες που σαρώθηκαν: {len(self.results)}\n"
                f"Συνολικό μέγεθος: {format_bytes(total)}\n\n"
                "Για να δημιουργηθεί η αναφορά, πάτησε Export Excel.",
            )

    def _export(self, show_popup: bool = True) -> None:
        if not self.results and not self.deletion_log:
            if show_popup:
                messagebox.showinfo("Export", "Δεν υπάρχουν αποτελέσματα για export.")
            return
        try:
            exported = export_results(self.results, self.output_dir_var.get().strip(), deletion_log=self.deletion_log)
            self.last_exported = exported
            self.status_var.set(
                f"Export Excel ολοκληρώθηκε | Excel: {exported['excel']}"
            )
            if show_popup:
                messagebox.showinfo(
                    "Export Excel ολοκληρώθηκε",
                    "Δημιουργήθηκε το αρχείο:\n\n"
                    f"Excel: {exported['excel']}",
                )
        except Exception as exc:
            self.status_var.set("Σφάλμα κατά το Excel export.")
            messagebox.showerror("Σφάλμα Excel export", str(exc))


def main() -> None:
    if is_windows() and not is_admin():
        root = tk.Tk()
        root.withdraw()
        try:
            messagebox.showerror(
                "Administrator required",
                "Το Storage Analyzer πρέπει να εκτελεστεί ως Administrator.\n\n"
                "Κάνε δεξί κλικ στο πρόγραμμα και επίλεξε Run as administrator.",
            )
        finally:
            root.destroy()
        sys.exit(1)

    app = StorageAnalyzerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
